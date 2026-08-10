import os
import re

from ...imkit import EagerPageSource
from .com import export_all_charts

from ...logging_config import get_logger, log_call

logger = get_logger(__name__)

_CHART_TYPE_LABEL = {"BarChart": "cột", "LineChart": "đường", "PieChart": "tròn"}
_EMU_PER_PX = 914400 / 96          # 96 DPI — cùng mốc quy đổi với DEF_CW/DEF_RH


@log_call
def _chart_title_text(chart):
    """Chữ tiêu đề chart (nếu có) — chỉ dùng cho placeholder khi không xuất
    được ảnh chart thật qua COM."""
    try:
        title = chart.title
        if title and title.tx and title.tx.rich:
            parts = [run.t for para in title.tx.rich.p for run in (para.r or []) if run.t]
            return " ".join(parts) or None
    except Exception:
        logger.debug("Could not read chart title text.", exc_info=True)
    return None


@log_call
def _chart_extent(chart, def_cw: int, def_rh: int):
    """Vị trí/kích thước chart theo cột-hàng (1-based, đơn vị pixel CHƯA nhân
    S — cùng thang với cw_list/rh_list) từ anchor đọc lại sau khi lưu file
    (OneCellAnchor/TwoCellAnchor có `_from`/`ext` tính bằng EMU). Dùng để (1)
    mở rộng canvas nếu chart nằm ngoài vùng dữ liệu, (2) biết ô để ghép ảnh."""
    anchor = getattr(chart, "anchor", None)
    frm = getattr(anchor, "_from", None)
    if frm is None:
        return None
    start_col, start_row = frm.col + 1, frm.row + 1
    ext = getattr(anchor, "ext", None)
    w_px, h_px = (ext.cx / _EMU_PER_PX, ext.cy / _EMU_PER_PX) if ext is not None else (480, 288)
    cols_needed = max(1, int(w_px / def_cw) + 1)
    rows_needed = max(1, int(h_px / def_rh) + 1)
    return start_col, start_row, cols_needed, rows_needed


@log_call
def excel_page_source(path: str) -> EagerPageSource:
    png_paths, tmp_dir = _render_excel_sheets(path)
    return EagerPageSource(png_paths, tmp_dir)


@log_call
def _render_excel_sheets(path: str) -> tuple[list[str], str]:
    import tempfile
    import openpyxl
    from openpyxl.utils import get_column_letter
    from PIL import Image, ImageDraw, ImageFont

    scale = 1.5; CHAR_PX = 7; PT_PX = 1.33
    ROW_HDR = 38; COL_HDR = 20; DEF_CW = 64; DEF_RH = 18
    GRID_CLR = "#C0C0C0"; HDR_BG = "#E8E8E8"; HDR_FG = "#666666"
    FORMULA_FG = "#6B6BC9"
    FONT_DIR = "C:/Windows/Fonts/"
    THEME = ["#FFFFFF","#000000","#E7E6E6","#44546A","#4472C4","#ED7D31",
             "#A5A5A5","#FFC000","#5B9BD5","#70AD47"]

    _fc: dict = {}

    def _fnt(size=11, bold=False):
        key = (int(size), bold)
        if key not in _fc:
            try:
                fn = "calibrib.ttf" if bold else "calibri.ttf"
                _fc[key] = ImageFont.truetype(FONT_DIR + fn, int(size * scale))
            except Exception:
                _fc[key] = ImageFont.load_default()
        return _fc[key]

    def _clr(color_obj, dflt="white"):
        try:
            if color_obj.type == "rgb":
                rgb = color_obj.rgb or ""
                if len(rgb) >= 6 and rgb != "00000000":
                    return "#" + rgb[-6:]
            elif color_obj.type == "theme":
                idx = color_obj.theme
                if 0 <= idx < len(THEME):
                    clr = THEME[idx]
                    tint = getattr(color_obj, "tint", 0) or 0
                    if tint > 0:
                        red, green, blue = (int(clr[pos:pos+2], 16) for pos in (1, 3, 5))
                        return f"#{int(red+(255-red)*tint):02X}{int(green+(255-green)*tint):02X}{int(blue+(255-blue)*tint):02X}"
                    elif tint < 0:
                        red, green, blue = (int(clr[pos:pos+2], 16) for pos in (1, 3, 5))
                        return f"#{int(red*(1+tint)):02X}{int(green*(1+tint)):02X}{int(blue*(1+tint)):02X}"
                    return clr
        except Exception:
            pass
        return dflt

    def _fmt_val(cell):
        value = cell.value
        if value is None:
            return ""
        fmt = (cell.number_format or "General").strip()
        if isinstance(value, (int, float)):
            if fmt in ("General", "@", ""):
                return str(int(value)) if float(value) == int(value) else str(value)
            if "%" in fmt:
                return f"{value * 100:.1f}%"
            dec = len([ch for ch in fmt.split(".")[-1] if ch in "0#"]) if "." in fmt else 0
            use_comma = "," in fmt
            if use_comma and dec == 0:
                return f"{int(value):,}"
            elif use_comma:
                return f"{value:,.{dec}f}"
            elif dec:
                return f"{value:.{dec}f}"
            return str(int(value)) if float(value) == int(value) else str(value)
        return str(value)

    wb = openpyxl.load_workbook(path, data_only=True)
    # Load riêng data_only=False để lấy chữ công thức khi openpyxl chưa có giá
    # trị cache (công thức do chính AI vừa ghi — chưa từng mở lại bằng Excel
    # thật nên không có kết quả tính sẵn, value ở bản data_only=True là None).
    wb_formulas = openpyxl.load_workbook(path, data_only=False)
    tmp_dir = tempfile.mkdtemp(prefix="docai_pages_")
    png_paths: list[str] = []

    has_charts = any(ws._charts for ws in wb.worksheets)
    chart_pngs: dict = {}
    if has_charts:
        chart_pngs = export_all_charts(path, tmp_dir)

    for ws in wb.worksheets:
        if not ws.max_row or not ws.max_column:
            continue
        if ws.max_row == 1 and ws.max_column == 1 and not ws.cell(1, 1).value:
            continue
        MR, MC = ws.max_row, ws.max_column
        formula_ws = wb_formulas[ws.title] if ws.title in wb_formulas.sheetnames else None

        chart_extents = []
        for ci, chart in enumerate(ws._charts):
            extent = _chart_extent(chart, DEF_CW, DEF_RH)
            if extent:
                start_col, start_row, cols_needed, rows_needed = extent
                chart_extents.append((ci, chart, start_col, start_row, cols_needed, rows_needed))
                MC = max(MC, start_col + cols_needed - 1)
                MR = max(MR, start_row + rows_needed - 1)

        cw_list = []
        for ci in range(1, MC + 1):
            col_dim = ws.column_dimensions.get(get_column_letter(ci))
            if col_dim and col_dim.hidden:     cw_list.append(0)
            elif col_dim and col_dim.width:    cw_list.append(max(4, int(col_dim.width * CHAR_PX)))
            else:                              cw_list.append(DEF_CW)

        rh_list = []
        for ri in range(1, MR + 1):
            row_dim = ws.row_dimensions.get(ri)
            if row_dim and row_dim.hidden:     rh_list.append(0)
            elif row_dim and row_dim.height:   rh_list.append(max(4, int(row_dim.height * PT_PX)))
            else:                              rh_list.append(DEF_RH)

        merged: dict = {}
        for mrng in ws.merged_cells.ranges:
            for row in range(mrng.min_row, mrng.max_row + 1):
                for col in range(mrng.min_col, mrng.max_col + 1):
                    merged[(row, col)] = mrng

        img_w = int((ROW_HDR + sum(cw_list)) * scale)
        img_h = int((COL_HDR + sum(rh_list)) * scale)
        img = Image.new("RGB", (img_w, img_h), "white")
        dw = ImageDraw.Draw(img)
        rh0 = int(ROW_HDR * scale)
        ch0 = int(COL_HDR * scale)

        dw.rectangle([0, 0, img_w - 1, ch0 - 1], fill=HDR_BG)
        dw.line([(0, ch0 - 1), (img_w - 1, ch0 - 1)], fill=GRID_CLR)
        dw.line([(rh0 - 1, 0), (rh0 - 1, ch0 - 1)], fill=GRID_CLR)
        x_pos = rh0
        for ci, pw in enumerate(cw_list):
            if pw == 0:
                continue
            px = int(pw * scale)
            ltr = get_column_letter(ci + 1)
            bb = dw.textbbox((0, 0), ltr, font=_fnt(8))
            tw, th = bb[2]-bb[0], bb[3]-bb[1]
            dw.text((x_pos + (px-tw)//2, (ch0-th)//2), ltr, fill=HDR_FG, font=_fnt(8))
            dw.line([(x_pos + px - 1, 0), (x_pos + px - 1, ch0 - 1)], fill=GRID_CLR)
            x_pos += px

        y_pos = ch0
        for ri, ph in enumerate(rh_list):
            if ph == 0:
                continue
            row_num = ri + 1
            rh_px = int(ph * scale)
            dw.rectangle([0, y_pos, rh0 - 1, y_pos + rh_px - 1], fill=HDR_BG)
            lbl = str(row_num)
            bb = dw.textbbox((0, 0), lbl, font=_fnt(8))
            tw2, th2 = bb[2]-bb[0], bb[3]-bb[1]
            dw.text((rh0//2 - tw2//2, y_pos + rh_px//2 - th2//2), lbl, fill=HDR_FG, font=_fnt(8))
            dw.line([(0, y_pos + rh_px - 1), (rh0 - 1, y_pos + rh_px - 1)], fill=GRID_CLR)
            dw.line([(rh0 - 1, y_pos), (rh0 - 1, y_pos + rh_px - 1)], fill=GRID_CLR)

            x_pos = rh0
            for ci, pw in enumerate(cw_list):
                if pw == 0:
                    continue
                col_num = ci + 1
                cw_px = int(pw * scale)
                cell = ws.cell(row_num, col_num)
                mrng = merged.get((row_num, col_num))
                slave = mrng and (mrng.min_row != row_num or mrng.min_col != col_num)

                if not slave:
                    dw2 = sum(int(cw_list[merge_col-1]*scale) for merge_col in range(mrng.min_col, mrng.max_col+1)) if mrng else cw_px
                    dh2 = sum(int(rh_list[merge_row-1]*scale) for merge_row in range(mrng.min_row, mrng.max_row+1)) if mrng else rh_px
                    bg = "white"
                    try:
                        if cell.fill and cell.fill.patternType == "solid":
                            bg = _clr(cell.fill.fgColor, "white")
                    except Exception:
                        pass
                    dw.rectangle([x_pos, y_pos, x_pos+dw2-1, y_pos+dh2-1], fill=bg)
                    text = _fmt_val(cell)
                    is_formula_text = False
                    if not text and formula_ws is not None:
                        fv = formula_ws.cell(row_num, col_num).value
                        if isinstance(fv, str) and fv.startswith("="):
                            text, is_formula_text = fv, True
                    if text:
                        try:
                            bold   = bool(cell.font.bold) if cell.font else False
                            fsize  = float(cell.font.size or 11) if cell.font else 11
                            fg     = _clr(cell.font.color, "#000000") if cell.font else "#000000"
                            halign = (cell.alignment.horizontal or "general") if cell.alignment else "general"
                        except Exception:
                            bold, fsize, fg, halign = False, 11, "#000000", "general"
                        if is_formula_text:
                            fg = FORMULA_FG
                        fnt = _fnt(fsize, bold)
                        pad = int(3 * scale)
                        try:
                            while len(text) > 1:
                                bb = dw.textbbox((0, 0), text, font=fnt)
                                if bb[2]-bb[0] <= dw2 - 2*pad:
                                    break
                                text = text[:-1]
                            bb = dw.textbbox((0, 0), text, font=fnt)
                            tw3 = bb[2]-bb[0]
                            th3 = bb[3]-bb[1]
                        except Exception:
                            tw3 = len(text)*7
                            th3 = 13
                        if halign in ("center", "centerContinuous", "distributed"):
                            tx = x_pos + (dw2 - tw3) // 2
                        elif halign == "right":
                            tx = x_pos + dw2 - tw3 - pad
                        else:
                            tx = x_pos + pad
                        dw.text((tx, y_pos + (dh2 - th3) // 2), text, fill=fg, font=fnt)

                    def _bl(side, x1, y1, x2, y2, _cell=cell):
                        try:
                            bs = getattr(_cell.border, side, None)
                            if bs and bs.border_style and bs.border_style != "none":
                                clr = _clr(bs.color, "#000000") if bs.color else "#000000"
                                width = 2 if bs.border_style in ("medium", "thick") else 1
                                dw.line([(x1, y1), (x2, y2)], fill=clr, width=int(width*scale))
                        except Exception:
                            pass
                    _bl("top",    x_pos,        y_pos,        x_pos+dw2-1, y_pos)
                    _bl("bottom", x_pos,        y_pos+dh2-1,  x_pos+dw2-1, y_pos+dh2-1)
                    _bl("left",   x_pos,        y_pos,        x_pos,        y_pos+dh2-1)
                    _bl("right",  x_pos+dw2-1,  y_pos,        x_pos+dw2-1, y_pos+dh2-1)

                dw.line([(x_pos, y_pos+rh_px-1), (x_pos+cw_px-1, y_pos+rh_px-1)], fill=GRID_CLR)
                dw.line([(x_pos+cw_px-1, y_pos), (x_pos+cw_px-1, y_pos+rh_px-1)], fill=GRID_CLR)
                x_pos += cw_px
            y_pos += rh_px

        for ci, chart, start_col, start_row, cols_needed, rows_needed in chart_extents:
            cx = int((ROW_HDR + sum(cw_list[:start_col - 1])) * scale)
            cy = int((COL_HDR + sum(rh_list[:start_row - 1])) * scale)
            cw_px = max(int(sum(cw_list[start_col - 1:start_col - 1 + cols_needed]) * scale), 60)
            ch_px = max(int(sum(rh_list[start_row - 1:start_row - 1 + rows_needed]) * scale), 60)
            png = chart_pngs.get((ws.title, ci))
            pasted = False
            if png and os.path.exists(png):
                try:
                    chart_img = Image.open(png).convert("RGB").resize((cw_px, ch_px))
                    img.paste(chart_img, (cx, cy))
                    pasted = True
                except Exception:
                    logger.warning("Failed to paste exported chart image %s — using placeholder.",
                                   png, exc_info=True)
                    pasted = False
            if not pasted:
                dw.rectangle([cx, cy, cx + cw_px - 1, cy + ch_px - 1], fill="#F2F2F2", outline=GRID_CLR)
                kind = _CHART_TYPE_LABEL.get(type(chart).__name__, "biểu đồ")
                title = _chart_title_text(chart)
                ty = cy + 8
                for line in ([f"[Biểu đồ {kind}]"] + ([title] if title else [])):
                    dw.text((cx + 8, ty), line, fill="#888888", font=_fnt(9))
                    ty += 16

        safe = re.sub(r'[^\w]', '_', ws.title)
        out = os.path.join(tmp_dir, f"sheet_{safe}.png")
        img.save(out, "PNG")
        png_paths.append(out)

    return png_paths, tmp_dir
