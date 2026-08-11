"""The set of Excel (.xlsx) edit operations for the edit-via-chat flow.

Two parts:
  · outline()/outline_text() — describes the spreadsheet's structure (sheets,
    data ranges, tables, charts) sent to the server as context for Claude.
  · apply_edits() — executes the list of edit commands Claude returns.

Cell/range addresses use A1 notation (e.g. "B2", "B2:D10"). Every command
(except add_sheet/delete_sheet/rename_sheet, which operate at the workbook
level) can include "sheet" (sheet name, blank = the active sheet).
"""
import copy
import re
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string, range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo

from ...logging_config import get_logger, log_call

logger = get_logger(__name__)


class ExcelOpError(Exception):
    """Invalid edit command — message shown to the user in Vietnamese."""


# Allowed operations — also the list the server describes to Claude.
OPS = (
    "set_cell",              # set a value/formula for 1 cell
    "fill_range",            # fill 2D data starting from 1 cell
    "append_row",            # add 1 full row at the end of the sheet
    "insert_row",            # insert an empty row
    "delete_row",            # delete a row
    "insert_column",         # insert an empty column
    "delete_column",         # delete a column
    "format_cells",          # font/size/color/fill/border/number/align for a cell or range
    "merge_cells",           # merge a range of cells
    "unmerge_cells",         # split a merged range of cells
    "create_table",          # turn a data range into an Excel Table
    "sort_range",            # sort a data range by 1+ columns
    "filter_range",          # filter: hide rows that don't match a condition on 1 column
    "clear_filter",          # clear the filter, show all rows again
    "create_chart",          # create a bar/line/pie chart
    "delete_chart",          # delete a chart
    "add_sheet",             # add a new sheet
    "delete_sheet",          # delete a sheet
    "rename_sheet",          # rename a sheet
    "freeze_panes",          # freeze rows/columns while scrolling
)

_CHART_CLASSES = {"bar": BarChart, "line": LineChart, "pie": PieChart}
_CHART_CLASSES_REV = {cls: name for name, cls in _CHART_CLASSES.items()}
_CHART_LABEL = {"bar": "cột", "line": "đường", "pie": "tròn"}

_ALIGN_H = {"left", "center", "right", "justify"}
_ALIGN_H_LABEL = {"left": "căn trái", "center": "căn giữa",
                  "right": "căn phải", "justify": "căn đều"}
_ALIGN_V = {"top", "center", "bottom"}
_ALIGN_V_LABEL = {"top": "căn trên", "center": "căn giữa dọc", "bottom": "căn dưới"}
_BORDER_SIDES = ("top", "bottom", "left", "right")

_MAX_SHEETS_DETAIL = 8
_MAX_ROWS_PREVIEW = 12
_MAX_COLS_PREVIEW = 18
_CELL_TRUNC = 40


# ── Shared helpers ────────────────────────────────────────────────────────────

@log_call
def _need(edit: dict, key: str):
    val = edit.get(key)
    if val is None or (isinstance(val, str) and not val.strip()):
        raise ExcelOpError(f"Lệnh «{edit.get('op')}» thiếu thông tin: {key}.")
    return val


@log_call
def _short(text, limit: int = 40) -> str:
    text = " ".join(str(text).split())
    return text if len(text) <= limit else text[:limit] + "…"


@log_call
def _sheet_at(wb, edit: dict):
    name = edit.get("sheet")
    if not name:
        return wb.active
    if name not in wb.sheetnames:
        raise ExcelOpError(
            f"Không có sheet «{name}». Các sheet hiện có: "
            + ", ".join(wb.sheetnames) + ".")
    return wb[name]


@log_call
def _parse_cell(value) -> tuple:
    try:
        col, row = coordinate_from_string(str(value).strip().upper())
        return col, row
    except Exception:
        raise ExcelOpError(f"Ô không hợp lệ: {value}.")


@log_call
def _range_bounds(edit: dict, key: str = "range"):
    value = _need(edit, key)
    try:
        min_col, min_row, max_col, max_row = range_boundaries(str(value).strip().upper())
    except Exception:
        logger.debug("range_boundaries() failed to parse %r", value, exc_info=True)
        min_col = None
    if min_col is None:
        raise ExcelOpError(f"Vùng không hợp lệ: {value}.")
    return min_col, min_row, max_col, max_row


@log_call
def _col_index_in_range(min_col: int, max_col: int, col_ref) -> int:
    """column in keys/filter_range: an int = 0-based offset within the range,
    a string = an absolute column letter (e.g. "C")."""
    if isinstance(col_ref, int):
        idx = min_col + col_ref
    else:
        col_ref_str = str(col_ref).strip()
        idx = min_col + int(col_ref_str) if col_ref_str.lstrip("-").isdigit() else column_index_from_string(col_ref_str.upper())
    if not min_col <= idx <= max_col:
        raise ExcelOpError(f"Cột «{col_ref}» nằm ngoài vùng đã cho.")
    return idx


@log_call
def _clean_argb(color) -> str:
    hex_color = str(color).strip().lstrip("#").upper()
    if len(hex_color) != 6 or any(ch not in "0123456789ABCDEF" for ch in hex_color):
        raise ExcelOpError(f"Mã màu không hợp lệ: {color}.")
    return "FF" + hex_color


# ── Read structure ────────────────────────────────────────────────────────

@log_call
def _cell_preview(value) -> str:
    if value is None:
        return ""
    text = " ".join(str(value).split())
    return text if len(text) <= _CELL_TRUNC else text[:_CELL_TRUNC] + "…"


@log_call
def _chart_title(chart):
    try:
        if chart.title and chart.title.tx and chart.title.tx.rich:
            parts = [run.t for para in chart.title.tx.rich.p for run in (para.r or []) if run.t]
            return " ".join(parts) or None
    except Exception:
        logger.debug("Could not read chart title from rich-text structure.", exc_info=True)
    return None


@log_call
def outline(path: str) -> dict:
    """The spreadsheet's structure to send to the server as context for Claude."""
    wb = openpyxl.load_workbook(path, data_only=False)
    active_title = wb.active.title if wb.active is not None else None
    sheets = []
    for si, ws in enumerate(wb.worksheets):
        info = {
            "name": ws.title,
            "active": ws.title == active_title,
            "max_row": ws.max_row or 0,
            "max_col": ws.max_column or 0,
            "freeze_panes": ws.freeze_panes,
            "tables": [{"name": name, "ref": ref} for name, ref in ws.tables.items()],
            "charts": [{"title": _chart_title(chart), "type": _CHART_CLASSES_REV.get(type(chart), type(chart).__name__)}
                       for chart in ws._charts],
            "detail": si < _MAX_SHEETS_DETAIL,
        }
        if info["detail"] and ws.max_row >= 1 and ws.max_column >= 1:
            n_cols = min(ws.max_column, _MAX_COLS_PREVIEW)
            info["header"] = [_cell_preview(ws.cell(1, col).value) for col in range(1, n_cols + 1)]
            info["cols_truncated"] = ws.max_column - n_cols
            n_rows = min(max(ws.max_row - 1, 0), _MAX_ROWS_PREVIEW)
            info["rows"] = [
                [_cell_preview(ws.cell(row, col).value) for col in range(1, n_cols + 1)]
                for row in range(2, 2 + n_rows)
            ]
            info["rows_truncated"] = max(0, ws.max_row - 1 - n_rows)
        sheets.append(info)
    return {"sheets": sheets, "sheet_count": len(wb.worksheets)}


@log_call
def outline_text(path: str) -> str:
    """A compact text-form outline for the prompt."""
    data = outline(path)
    lines = [f"Bảng tính có {data['sheet_count']} sheet."]
    for info in data["sheets"]:
        tag = " (đang mở)" if info["active"] else ""
        lines.append(f"\n[Sheet «{info['name']}»]{tag} {info['max_row']} dòng × {info['max_col']} cột.")
        if info.get("freeze_panes"):
            lines.append(f"Freeze: cố định trước ô {info['freeze_panes']}.")
        for table in info["tables"]:
            lines.append(f"Bảng: «{table['name']}» {table['ref']}")
        for chart in info["charts"]:
            title = f" «{chart['title']}»" if chart.get("title") else ""
            lines.append(f"Biểu đồ ({chart['type']}){title}")
        if not info["detail"]:
            lines.append("(sheet còn lại — không hiển thị chi tiết)")
            continue
        if "header" not in info:
            lines.append("(sheet trống)")
            continue
        header = " | ".join(cell for cell in info["header"] if cell) or "(trống)"
        lines.append(f"Hàng đầu: {header}")
        for row_number, row in enumerate(info["rows"], start=2):
            lines.append(f"[{row_number}] " + " | ".join(row))
        if info["rows_truncated"]:
            lines.append(f"(còn {info['rows_truncated']} dòng nữa không hiển thị)")
        if info["cols_truncated"]:
            lines.append(f"(+{info['cols_truncated']} cột nữa)")
    return "\n".join(lines)


# ── Execute edit commands ────────────────────────────────────────────────────

@log_call
def apply_edits(path: str, edits: list[dict]) -> list[str]:
    """Apply the edit commands to the file in order, save once. Returns a
    description of each command."""
    if not edits:
        return []

    wb = openpyxl.load_workbook(path)
    notes: list[str] = []
    for edit in edits:
        op = (edit.get("op") or "").strip()
        if op not in OPS:
            raise ExcelOpError(f"Thao tác chưa hỗ trợ: «{op}».")
        notes.append(_run_op(wb, op, edit))

    try:
        wb.save(path)
    except PermissionError:
        raise ExcelOpError(
            f"Không ghi được «{Path(path).name}» — file đang mở trong Excel, "
            "hãy đóng lại rồi thử lại.")
    return notes


@log_call
def _run_op(wb, op: str, edit: dict) -> str:
    if op == "add_sheet":
        return _add_sheet(wb, edit)
    if op == "delete_sheet":
        return _delete_sheet(wb, edit)
    if op == "rename_sheet":
        return _rename_sheet(wb, edit)

    ws = _sheet_at(wb, edit)

    if op == "set_cell":
        return _set_cell(ws, edit)
    if op == "fill_range":
        return _fill_range(ws, edit)
    if op == "append_row":
        return _append_row(ws, edit)
    if op == "insert_row":
        return _insert_row(ws, edit)
    if op == "delete_row":
        return _delete_row(ws, edit)
    if op == "insert_column":
        return _insert_column(ws, edit)
    if op == "delete_column":
        return _delete_column(ws, edit)
    if op == "format_cells":
        return _format_cells(ws, edit)
    if op == "merge_cells":
        return _merge_cells(ws, edit)
    if op == "unmerge_cells":
        return _unmerge_cells(ws, edit)
    if op == "create_table":
        return _create_table(ws, wb, edit)
    if op == "sort_range":
        return _sort_range(ws, edit)
    if op == "filter_range":
        return _filter_range(ws, edit)
    if op == "clear_filter":
        return _clear_filter(ws, edit)
    if op == "create_chart":
        return _create_chart(ws, edit)
    if op == "delete_chart":
        return _delete_chart(ws, edit)

    # freeze_panes — the last remaining op in OPS.
    return _freeze_panes(ws, edit)


# ── Cells & ranges — basic content ───────────────────────────────────────────

@log_call
def _set_cell(ws, edit: dict) -> str:
    cell_ref = _need(edit, "cell")
    col, row = _parse_cell(cell_ref)
    value = edit.get("value")
    if value is None:
        raise ExcelOpError("Lệnh «set_cell» thiếu thông tin: value.")
    ws[f"{col}{row}"] = value
    kind = "công thức" if isinstance(value, str) and value.strip().startswith("=") else "giá trị"
    return f"đặt {kind} «{_short(value)}» vào ô {col}{row} ở sheet «{ws.title}»"


@log_call
def _fill_range(ws, edit: dict) -> str:
    start = _need(edit, "start_cell")
    col0, row0 = _parse_cell(start)
    col0_idx = column_index_from_string(col0)
    data = _need(edit, "data")
    if not isinstance(data, list) or not all(isinstance(row_vals, list) for row_vals in data):
        raise ExcelOpError("«fill_range» cần data là mảng 2 chiều (mảng các hàng).")
    for row_offset, row_vals in enumerate(data):
        for col_offset, val in enumerate(row_vals):
            ws.cell(row=row0 + row_offset, column=col0_idx + col_offset, value=val)
    rows = len(data)
    cols = max((len(row_vals) for row_vals in data), default=0)
    return f"điền dữ liệu {rows}×{cols} từ ô {start} ở sheet «{ws.title}»"


@log_call
def _append_row(ws, edit: dict) -> str:
    values = edit.get("values")
    if not isinstance(values, list) or not values:
        raise ExcelOpError("«append_row» cần values là danh sách giá trị.")
    ws.append(values)
    preview = " | ".join(_short(val, 20) for val in values)
    return f"thêm dòng mới «{preview}» vào cuối sheet «{ws.title}»"


@log_call
def _insert_row(ws, edit: dict) -> str:
    index = int(_need(edit, "index"))
    count = int(edit.get("count") or 1)
    if index < 1:
        raise ExcelOpError("Vị trí chèn hàng phải từ 1 trở lên.")
    ws.insert_rows(index, amount=count)
    return f"chèn {count} hàng mới tại vị trí {index} ở sheet «{ws.title}»"


@log_call
def _delete_row(ws, edit: dict) -> str:
    index = int(_need(edit, "index"))
    count = int(edit.get("count") or 1)
    if index < 1 or index > ws.max_row:
        raise ExcelOpError(f"Sheet «{ws.title}» chỉ có {ws.max_row} hàng, không có hàng {index}.")
    ws.delete_rows(index, amount=count)
    return f"xóa {count} hàng từ vị trí {index} ở sheet «{ws.title}»"


@log_call
def _insert_column(ws, edit: dict) -> str:
    index = _col_index_value(_need(edit, "index"))
    count = int(edit.get("count") or 1)
    if index < 1:
        raise ExcelOpError("Vị trí chèn cột phải từ 1 trở lên.")
    ws.insert_cols(index, amount=count)
    return f"chèn {count} cột mới tại vị trí {get_column_letter(index)} ở sheet «{ws.title}»"


@log_call
def _delete_column(ws, edit: dict) -> str:
    index = _col_index_value(_need(edit, "index"))
    count = int(edit.get("count") or 1)
    if index < 1 or index > ws.max_column:
        raise ExcelOpError(
            f"Sheet «{ws.title}» chỉ có {ws.max_column} cột, "
            f"không có cột {get_column_letter(index)}.")
    ws.delete_cols(index, amount=count)
    return f"xóa {count} cột từ vị trí {get_column_letter(index)} ở sheet «{ws.title}»"


@log_call
def _col_index_value(value) -> int:
    """index for insert_column/delete_column: an int (1-based) or a column letter."""
    if isinstance(value, int):
        return value
    value_str = str(value).strip()
    if value_str.isdigit():
        return int(value_str)
    try:
        return column_index_from_string(value_str.upper())
    except Exception:
        raise ExcelOpError(f"Cột không hợp lệ: {value}.")


# ── Formatting ───────────────────────────────────────────────────────────────

@log_call
def _format_cells(ws, edit: dict) -> str:
    min_col, min_row, max_col, max_row = _range_bounds(edit)
    changes = []

    font_kwargs = {}
    if edit.get("font_name"):
        font_kwargs["name"] = str(edit["font_name"])
        changes.append(f"font «{edit['font_name']}»")
    if edit.get("font_size"):
        font_kwargs["size"] = float(edit["font_size"])
        changes.append(f"cỡ chữ {float(edit['font_size']):g}pt")
    if edit.get("font_color"):
        font_kwargs["color"] = _clean_argb(edit["font_color"])
        changes.append(f"màu chữ #{str(edit['font_color']).lstrip('#').upper()}")
    if edit.get("bold") is not None:
        font_kwargs["bold"] = bool(edit["bold"])
        changes.append("in đậm" if edit["bold"] else "bỏ in đậm")
    if edit.get("italic") is not None:
        font_kwargs["italic"] = bool(edit["italic"])
        changes.append("in nghiêng" if edit["italic"] else "bỏ in nghiêng")
    if edit.get("underline") is not None:
        font_kwargs["underline"] = "single" if edit["underline"] else "none"
        changes.append("gạch chân" if edit["underline"] else "bỏ gạch chân")

    fill = None
    if edit.get("fill_color"):
        fill = PatternFill(fill_type="solid", fgColor=_clean_argb(edit["fill_color"]))
        changes.append(f"màu nền #{str(edit['fill_color']).lstrip('#').upper()}")

    number_format = edit.get("number_format")
    if number_format:
        changes.append(f"định dạng số «{number_format}»")

    align_kwargs = {}
    if edit.get("align"):
        align = str(edit["align"]).lower()
        if align not in _ALIGN_H:
            raise ExcelOpError(f"Kiểu căn ngang chưa hỗ trợ: {align}.")
        align_kwargs["horizontal"] = align
        changes.append(_ALIGN_H_LABEL[align])
    if edit.get("valign"):
        valign = str(edit["valign"]).lower()
        if valign not in _ALIGN_V:
            raise ExcelOpError(f"Kiểu căn dọc chưa hỗ trợ: {valign}.")
        align_kwargs["vertical"] = valign
        changes.append(_ALIGN_V_LABEL[valign])
    if edit.get("wrap_text") is not None:
        align_kwargs["wrap_text"] = bool(edit["wrap_text"])
        changes.append("xuống dòng tự động" if edit["wrap_text"] else "bỏ xuống dòng tự động")

    border_side = None
    border_sides = []
    if edit.get("border"):
        color = _clean_argb(edit.get("border_color") or "000000")
        weight = "medium" if int(edit.get("border_size") or 1) >= 2 else "thin"
        border_side = Side(style=weight, color=color)
        border_sides = edit.get("border_sides") or list(_BORDER_SIDES)
        for side in border_sides:
            if side not in _BORDER_SIDES:
                raise ExcelOpError(f"border_sides không hợp lệ: {side}.")
        changes.append("viền " + "/".join(border_sides))

    if not changes:
        raise ExcelOpError("Lệnh «format_cells» không nêu định dạng nào.")

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if font_kwargs:
                base = cell.font
                cell.font = Font(
                    name=font_kwargs.get("name", base.name),
                    size=font_kwargs.get("size", base.size),
                    bold=font_kwargs.get("bold", base.bold),
                    italic=font_kwargs.get("italic", base.italic),
                    underline=font_kwargs.get("underline", base.underline),
                    color=font_kwargs.get("color", base.color),
                )
            if fill:
                cell.fill = fill
            if number_format:
                cell.number_format = str(number_format)
            if align_kwargs:
                base_a = cell.alignment
                cell.alignment = Alignment(
                    horizontal=align_kwargs.get("horizontal", base_a.horizontal),
                    vertical=align_kwargs.get("vertical", base_a.vertical),
                    wrap_text=align_kwargs.get("wrap_text", base_a.wrap_text),
                )
            if border_side:
                existing_border = cell.border
                sides = {"top": existing_border.top, "bottom": existing_border.bottom,
                         "left": existing_border.left, "right": existing_border.right}
                for side in border_sides:
                    sides[side] = border_side
                cell.border = Border(**sides)

    if (min_col, min_row) == (max_col, max_row):
        scope = f"{get_column_letter(min_col)}{min_row}"
    else:
        scope = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    return f"định dạng {scope} ở sheet «{ws.title}»: {', '.join(changes)}"


@log_call
def _merge_cells(ws, edit: dict) -> str:
    rng = _need(edit, "range")
    try:
        ws.merge_cells(str(rng))
    except Exception as exc:  # noqa: BLE001 — vùng chồng lấn ô đã gộp khác...
        raise ExcelOpError(f"Không gộp được ô: {exc}")
    return f"gộp ô {rng} ở sheet «{ws.title}»"


@log_call
def _unmerge_cells(ws, edit: dict) -> str:
    rng = _need(edit, "range")
    try:
        ws.unmerge_cells(str(rng))
    except Exception as exc:  # noqa: BLE001
        raise ExcelOpError(f"Không tách được ô: {exc}")
    return f"tách ô đã gộp {rng} ở sheet «{ws.title}»"


# ── Tables, sort, filter ──────────────────────────────────────────────────────

@log_call
def _sanitize_table_name(name, wb) -> str:
    name = re.sub(r"[^A-Za-z0-9_]", "_", str(name).strip())
    if not name or not name[0].isalpha():
        name = "Table_" + name if name else "Table1"
    existing = {table_name for ws in wb.worksheets for table_name in ws.tables.keys()}
    base, suffix = name, 1
    while name in existing:
        suffix += 1
        name = f"{base}_{suffix}"
    return name


@log_call
def _create_table(ws, wb, edit: dict) -> str:
    rng = _need(edit, "range")
    name = _sanitize_table_name(edit.get("name") or f"Table_{ws.title}", wb)
    style = edit.get("style") or "TableStyleMedium2"
    table = Table(displayName=name, ref=str(rng))
    table.tableStyleInfo = TableStyleInfo(
        name=style, showRowStripes=True, showFirstColumn=False,
        showLastColumn=False, showColumnStripes=False)
    try:
        ws.add_table(table)
    except Exception as exc:  # noqa: BLE001 — vùng trùng bảng khác, tên trùng...
        raise ExcelOpError(f"Không tạo được bảng: {exc}")
    return f"chuyển vùng {rng} thành bảng «{name}» (có header + lọc) ở sheet «{ws.title}»"


@log_call
def _cell_overlaps_merge(ws, min_row, max_row, min_col, max_col) -> bool:
    for mrng in ws.merged_cells.ranges:
        if not (mrng.max_row < min_row or mrng.min_row > max_row
                or mrng.max_col < min_col or mrng.min_col > max_col):
            return True
    return False


@log_call
def _sort_key(value):
    if value is None:
        return (0, "")
    if isinstance(value, (int, float)):
        return (1, value)
    return (2, str(value).lower())


@log_call
def _sort_range(ws, edit: dict) -> str:
    min_col, min_row, max_col, max_row = _range_bounds(edit)
    has_header = True if edit.get("has_header") is None else bool(edit["has_header"])
    data_min_row = min_row + 1 if has_header else min_row
    if data_min_row > max_row:
        raise ExcelOpError("Vùng sắp xếp không có dữ liệu.")

    if _cell_overlaps_merge(ws, data_min_row, max_row, min_col, max_col):
        raise ExcelOpError("Vùng sắp xếp có ô đã gộp — hãy tách ô trước khi sắp xếp.")

    for row in range(data_min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell_value = ws.cell(row=row, column=col).value
            if isinstance(cell_value, str) and cell_value.startswith("="):
                raise ExcelOpError(
                    "Vùng sắp xếp có công thức — hãy chuyển thành giá trị trước "
                    "(bản này chưa tự tính lại tham chiếu công thức khi đổi vị trí hàng).")

    keys = edit.get("keys")
    if not isinstance(keys, list) or not keys:
        raise ExcelOpError("«sort_range» cần keys là danh sách {column, order}.")

    key_specs = []
    for key_spec in keys:
        col_ref = key_spec.get("column") if isinstance(key_spec, dict) else None
        if col_ref is None:
            raise ExcelOpError("Mỗi phần tử keys cần có «column».")
        idx = _col_index_in_range(min_col, max_col, col_ref)
        order = str((key_spec.get("order") if isinstance(key_spec, dict) else None) or "asc").lower()
        if order not in ("asc", "desc"):
            raise ExcelOpError('order phải là "asc" hoặc "desc".')
        key_specs.append((idx, order == "desc"))

    rows_snapshot = []
    for row in range(data_min_row, max_row + 1):
        cells = []
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cells.append({
                "value": cell.value, "number_format": cell.number_format,
                "font": copy.copy(cell.font), "fill": copy.copy(cell.fill),
                "border": copy.copy(cell.border), "alignment": copy.copy(cell.alignment),
            })
        rows_snapshot.append(cells)

    for idx, desc in reversed(key_specs):
        offset = idx - min_col
        rows_snapshot.sort(key=lambda row_cells, off=offset: _sort_key(row_cells[off]["value"]), reverse=desc)

    for row_offset, cells in enumerate(rows_snapshot):
        row = data_min_row + row_offset
        for col_offset, cell_data in enumerate(cells):
            col = min_col + col_offset
            cell = ws.cell(row=row, column=col)
            cell.value = cell_data["value"]
            cell.number_format = cell_data["number_format"]
            cell.font = cell_data["font"]
            cell.fill = cell_data["fill"]
            cell.border = cell_data["border"]
            cell.alignment = cell_data["alignment"]

    rng = edit.get("range")
    desc_txt = ", ".join(
        f"{get_column_letter(idx)} ({'giảm' if desc else 'tăng'} dần)" for idx, desc in key_specs)
    return f"sắp xếp {rng} ở sheet «{ws.title}» theo {desc_txt}"


@log_call
def _norm_text(value) -> str:
    return str(value).strip().lower() if value is not None else ""


@log_call
def _as_number(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


@log_call
def _filter_match(condition: str, value, target) -> bool:
    if condition == "equals":
        return _norm_text(value) == _norm_text(target)
    if condition == "not_equals":
        return _norm_text(value) != _norm_text(target)
    if condition == "contains":
        return _norm_text(target) in _norm_text(value)
    if condition == "not_contains":
        return _norm_text(target) not in _norm_text(value)
    if condition == "is_empty":
        return value is None or str(value).strip() == ""
    if condition == "is_not_empty":
        return not (value is None or str(value).strip() == "")
    num = _as_number(value)
    target_num = _as_number(target)
    if num is None or target_num is None:
        return False
    if condition == "gt":
        return num > target_num
    if condition == "gte":
        return num >= target_num
    if condition == "lt":
        return num < target_num
    if condition == "lte":
        return num <= target_num
    raise ExcelOpError(f"Điều kiện lọc chưa hỗ trợ: {condition}.")


_FILTER_CONDITIONS = {
    "equals", "not_equals", "contains", "not_contains",
    "gt", "gte", "lt", "lte", "is_empty", "is_not_empty",
}


@log_call
def _filter_range(ws, edit: dict) -> str:
    min_col, min_row, max_col, max_row = _range_bounds(edit)
    if max_row <= min_row:
        raise ExcelOpError("Vùng lọc không có dữ liệu (chỉ có hàng tiêu đề).")
    col_ref = _need(edit, "column")
    col_idx = _col_index_in_range(min_col, max_col, col_ref)
    condition = _need(edit, "condition")
    if condition not in _FILTER_CONDITIONS:
        raise ExcelOpError(f"Điều kiện lọc chưa hỗ trợ: {condition}.")
    target = edit.get("value")
    if condition not in ("is_empty", "is_not_empty") and target is None:
        raise ExcelOpError(f"Điều kiện «{condition}» cần «value».")

    hidden_count = 0
    for row in range(min_row + 1, max_row + 1):
        cell_value = ws.cell(row=row, column=col_idx).value
        if not _filter_match(condition, cell_value, target):
            ws.row_dimensions[row].hidden = True
            hidden_count += 1

    rng = edit.get("range")
    ws.auto_filter.ref = str(rng)
    return (f"lọc {rng} ở sheet «{ws.title}»: cột {get_column_letter(col_idx)} "
            f"{condition} «{target}» (ẩn {hidden_count} dòng không khớp)")


@log_call
def _clear_filter(ws, edit: dict) -> str:
    for row_dim in ws.row_dimensions.values():
        row_dim.hidden = False
    ws.auto_filter.ref = None
    return f"bỏ lọc, hiện lại toàn bộ dòng ở sheet «{ws.title}»"


# ── Charts ───────────────────────────────────────────────────────────────────

@log_call
def _create_chart(ws, edit: dict) -> str:
    chart_type = str(_need(edit, "chart_type")).lower()
    cls = _CHART_CLASSES.get(chart_type)
    if cls is None:
        raise ExcelOpError('chart_type phải là "bar", "line" hoặc "pie".')

    data_range = _need(edit, "data_range")
    min_col, min_row, max_col, max_row = _range_bounds(edit, "data_range")
    series_by = str(edit.get("series_by") or "columns").lower()

    chart = cls()
    data_ref = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
    chart.add_data(data_ref, titles_from_data=True, from_rows=(series_by == "rows"))

    cats_range = edit.get("categories_range")
    if cats_range:
        c_min_col, c_min_row, c_max_col, c_max_row = range_boundaries(str(cats_range).strip().upper())
        cats_ref = Reference(ws, min_col=c_min_col, min_row=c_min_row, max_col=c_max_col, max_row=c_max_row)
        chart.set_categories(cats_ref)

    title = edit.get("title")
    if title:
        chart.title = str(title)

    anchor = edit.get("anchor") or "H2"
    ws.add_chart(chart, str(anchor))

    label = _CHART_LABEL[chart_type]
    title_txt = f" «{title}»" if title else ""
    return f"tạo biểu đồ {label}{title_txt} từ {data_range} ở sheet «{ws.title}»"


@log_call
def _delete_chart(ws, edit: dict) -> str:
    charts = ws._charts
    if not charts:
        raise ExcelOpError(f"Sheet «{ws.title}» không có biểu đồ nào.")
    index = edit.get("index")
    index = int(index) if index is not None else len(charts) - 1
    if not 0 <= index < len(charts):
        raise ExcelOpError(f"Sheet «{ws.title}» chỉ có {len(charts)} biểu đồ, không có biểu đồ {index}.")
    del charts[index]
    return f"xóa biểu đồ {index} ở sheet «{ws.title}»"


# ── Sheet & freeze pane ──────────────────────────────────────────────────────

@log_call
def _add_sheet(wb, edit: dict) -> str:
    name = _need(edit, "name")
    if name in wb.sheetnames:
        raise ExcelOpError(f"Sheet «{name}» đã tồn tại.")
    index = edit.get("index")
    ws = wb.create_sheet(title=str(name), index=int(index) if index is not None else None)
    return f"thêm sheet mới «{ws.title}»"


@log_call
def _delete_sheet(wb, edit: dict) -> str:
    name = _need(edit, "sheet")
    if name not in wb.sheetnames:
        raise ExcelOpError(f"Không có sheet «{name}».")
    if len(wb.sheetnames) == 1:
        raise ExcelOpError("Không thể xóa sheet cuối cùng của workbook.")
    del wb[name]
    return f"xóa sheet «{name}»"


@log_call
def _rename_sheet(wb, edit: dict) -> str:
    name = _need(edit, "sheet")
    new_name = _need(edit, "new_name")
    if name not in wb.sheetnames:
        raise ExcelOpError(f"Không có sheet «{name}».")
    if new_name in wb.sheetnames:
        raise ExcelOpError(f"Đã có sheet tên «{new_name}».")
    ws = wb[name]
    try:
        ws.title = str(new_name)
    except Exception as exc:  # noqa: BLE001 — tên chứa ký tự cấm, quá dài...
        raise ExcelOpError(f"Tên sheet không hợp lệ: {exc}")
    return f"đổi tên sheet «{name}» thành «{ws.title}»"


@log_call
def _freeze_panes(ws, edit: dict) -> str:
    if edit.get("unfreeze"):
        ws.freeze_panes = None
        return f"bỏ cố định vùng ở sheet «{ws.title}»"
    cell_ref = _need(edit, "cell")
    col, row = _parse_cell(cell_ref)
    ws.freeze_panes = f"{col}{row}"
    return f"cố định vùng trước ô {col}{row} ở sheet «{ws.title}»"
