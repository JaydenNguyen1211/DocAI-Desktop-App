"""Compute payroll along with social/health/unemployment insurance +
personal income tax from a timesheet.

Deliberately made DETERMINISTIC — never let Claude compute payroll/
insurance/tax figures itself via chat, per the principle in the planning
docs: "DocAI acts as an aggregation aid, leaving the user or accountant to
confirm the final figures" (Tax/Social-Insurance business area — high risk
if miscalculated).

Flow: the user opens a "timesheet" Excel file in DocAI then types a command
like "tính lương" ("compute payroll"), "tính BHXH"… → `detect_payroll_intent()`
recognizes it locally (no AI call spent) → `compute_payroll_file()` reads the
file, computes using the parameters in `payroll_rates.py`, writes out a NEW
"Payroll" file (the original timesheet file is untouched) — same staging +
"Click to preview & save" mechanism as `modules.common.doc_set`.

Column convention in the input "timesheet" file (row 1 = column headers,
read from the active sheet). Column name recognition is case/diacritic
insensitive, accepts a few common spellings — see `_COLUMN_ALIASES`. At
minimum "Full name" and "Base salary" are required; the remaining columns
fall back to reasonable defaults if missing.
"""
import os
import re
import tempfile
import time
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

from ...account import api_client
from ...account.config import load_config, save_config
from .payroll_rates import (
    DEFAULT_PAYROLL_RATES, InsuranceRates, PayrollRates, TaxBracket, normalize_region,
)

from ...logging_config import get_logger, log_call
from ...strings import Payroll as S

logger = get_logger(__name__)


class PayrollError(api_client.ApiError):
    """Couldn't compute payroll — message shown to the user in Vietnamese.

    Inherits from `ApiError` (like `DocSetError`/`ConvertError`) so
    `CallWorker` emits the right message via `err`, without wrapping it in
    an "Unknown error:" prefix — see `app/thread_worker.py`."""

    @log_call
    def __init__(self, message: str):
        super().__init__(message, "payroll_failed")


# ── Local chat command recognition (no AI call spent) ───────────────────────

_INTENT_RE = re.compile(
    r"tính\s*(lại)?\s*lương|bảng\s*lương|tính\s*bhxh|tính\s*bảo\s*hiểm|"
    r"\bbhxh\b|\bbhyt\b|\bbhtn\b|gross\s*[-→>]*\s*net|trích\s*bảo\s*hiểm",
    re.IGNORECASE)


@log_call
def detect_payroll_intent(text: str) -> bool:
    return bool(_INTENT_RE.search(text or ""))


# ── Read the timesheet ───────────────────────────────────────────────────────

_COLUMN_ALIASES = {
    "name": {"ho ten", "ho va ten", "ten nhan vien", "nhan vien", "ho ten nhan vien"},
    "basic_salary": {"luong co ban", "luong dong bhxh", "muc luong", "luong",
                      "muc luong dong bhxh"},
    "standard_days": {"ngay cong chuan", "cong chuan", "so ngay cong chuan"},
    "actual_days": {"ngay cong thuc te", "cong thuc te", "ngay cong", "so ngay cong"},
    "dependents": {"so nguoi phu thuoc", "nguoi phu thuoc",
                   "so nguoi phu thuoc tinh thue"},
    "region": {"vung", "khu vuc", "vung luong toi thieu"},
    "allowance": {"phu cap", "phu cap khac", "cac khoan phu cap"},
}
_REQUIRED_FIELDS = ("name", "basic_salary")
_COLUMN_LABEL_VN = {
    "name": "Họ tên", "basic_salary": "Lương cơ bản (lương đóng BHXH)",
    "standard_days": "Ngày công chuẩn", "actual_days": "Ngày công thực tế",
    "dependents": "Số người phụ thuộc", "region": "Vùng", "allowance": "Phụ cấp",
}


@log_call
def _strip_accents(text: str) -> str:
    norm = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in norm if not unicodedata.combining(ch)).replace("đ", "d").replace("Đ", "D")


@log_call
def _norm_header(text) -> str:
    return _strip_accents(str(text or "").strip().lower())


@log_call
def _match_columns(header_row: list) -> dict[str, int]:
    found: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        norm = _norm_header(raw)
        if not norm:
            continue
        for field_name, aliases in _COLUMN_ALIASES.items():
            if field_name not in found and norm in aliases:
                found[field_name] = idx
    missing = [f for f in _REQUIRED_FIELDS if f not in found]
    if missing:
        expected = "\n".join(f"  · {_COLUMN_LABEL_VN[f]}" for f in _COLUMN_ALIASES)
        names = ", ".join(_COLUMN_LABEL_VN[f] for f in missing)
        raise PayrollError(S.MISSING_COLUMNS.format(names=names, expected=expected))
    return found


@log_call
def _to_number(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


@dataclass
class EmployeeInput:
    name: str
    basic_salary: float
    standard_days: float = 26
    actual_days: float = 26
    dependents: int = 0
    region: str = "I"
    allowance: float = 0


@log_call
def read_attendance_table(path: str) -> list[EmployeeInput]:
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:  # noqa: BLE001 — all file-open errors collapse to one message
        raise PayrollError(S.READ_FAILED.format(name=Path(path).name, error=exc))

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise PayrollError(S.FILE_EMPTY.format(name=Path(path).name))

    cols = _match_columns(list(rows[0]))
    employees: list[EmployeeInput] = []
    for row in rows[1:]:
        if cols["name"] >= len(row):
            continue
        name = str(row[cols["name"]] or "").strip()
        if not name:
            continue

        def get(field_name: str, default=0):
            idx = cols.get(field_name)
            if idx is None or idx >= len(row):
                return default
            return row[idx]

        standard_days = _to_number(get("standard_days", 26)) or 26
        actual_days = _to_number(get("actual_days"))
        if not cols.get("actual_days"):
            actual_days = standard_days  # no "actual days worked" column → assume full attendance

        employees.append(EmployeeInput(
            name=name,
            basic_salary=_to_number(get("basic_salary")),
            standard_days=standard_days,
            actual_days=actual_days,
            dependents=int(_to_number(get("dependents", 0))),
            region=normalize_region(get("region", "I")),
            allowance=_to_number(get("allowance", 0)),
        ))

    if not employees:
        raise PayrollError(S.NO_VALID_EMPLOYEES.format(name=Path(path).name))
    return employees


# ── Compute progressive personal income tax ──────────────────────────────────

@log_call
def compute_progressive_tax(taxable_income: float,
                            brackets: list[TaxBracket] | None = None) -> int:
    brackets = brackets if brackets is not None else DEFAULT_PAYROLL_RATES.tax_brackets
    if taxable_income <= 0:
        return 0
    tax = 0.0
    prev_threshold = 0.0
    for bracket in brackets:
        if taxable_income > bracket.upto:
            tax += (bracket.upto - prev_threshold) * bracket.rate
            prev_threshold = bracket.upto
        else:
            tax += (taxable_income - prev_threshold) * bracket.rate
            break
    return round(tax)


# ── Compute payroll for 1 employee ───────────────────────────────────────────

@dataclass
class EmployeeResult:
    name: str
    region: str
    basic_salary: float
    actual_days: float
    standard_days: float
    allowance: float
    gross_income: float          # salary by days worked + allowance, before insurance deductions
    bhxh_nld: int
    bhyt_nld: int
    bhtn_nld: int
    total_nld: int                # total employee-side self-withheld amount
    dependents: int
    family_deduction: int
    taxable_income: int           # taxable income (after family deduction)
    pit: int                      # personal income tax
    net_income: int               # take-home pay
    bhxh_dn: int
    bhyt_dn: int
    bhtn_dn: int
    kpcd_dn: int
    total_dn_insurance: int
    total_dn_cost: int            # total employer cost for this employee (gross + employer-side insurance)


@log_call
def compute_employee(emp: EmployeeInput,
                     rates: PayrollRates = DEFAULT_PAYROLL_RATES) -> EmployeeResult:
    if emp.basic_salary <= 0:
        raise PayrollError(S.SALARY_MUST_BE_POSITIVE.format(name=emp.name))
    if emp.standard_days <= 0:
        raise PayrollError(S.STANDARD_DAYS_MUST_BE_POSITIVE.format(name=emp.name))

    ratio = min(1.0, emp.actual_days / emp.standard_days)
    luong_theo_cong = emp.basic_salary * ratio
    gross_income = luong_theo_cong + emp.allowance

    ins = rates.insurance
    # Social/health/unemployment insurance contribution is based on the
    # contracted salary (basic_salary), not actual days worked — a
    # simplified convention for the MVP, doesn't handle a full unpaid month
    # / maternity leave (Phase 2).
    muc_dong_bhxh = min(emp.basic_salary, rates.bhxh_bhyt_tran())
    muc_dong_bhtn = min(emp.basic_salary, rates.bhtn_tran(emp.region))

    bhxh_nld = round(muc_dong_bhxh * ins.bhxh_nld)
    bhyt_nld = round(muc_dong_bhxh * ins.bhyt_nld)
    bhtn_nld = round(muc_dong_bhtn * ins.bhtn_nld)
    total_nld = bhxh_nld + bhyt_nld + bhtn_nld

    family_deduction = round(
        rates.giam_tru_ban_than + emp.dependents * rates.giam_tru_nguoi_phu_thuoc)
    thu_nhap_chiu_thue = gross_income - total_nld
    taxable_income = max(0, round(thu_nhap_chiu_thue - family_deduction))
    pit = compute_progressive_tax(taxable_income, rates.tax_brackets)
    net_income = round(gross_income - total_nld - pit)

    bhxh_dn = round(muc_dong_bhxh * ins.bhxh_dn)
    bhyt_dn = round(muc_dong_bhxh * ins.bhyt_dn)
    bhtn_dn = round(muc_dong_bhtn * ins.bhtn_dn)
    kpcd_dn = round(muc_dong_bhxh * ins.kpcd_dn)
    total_dn_insurance = bhxh_dn + bhyt_dn + bhtn_dn + kpcd_dn
    total_dn_cost = round(gross_income + total_dn_insurance)

    return EmployeeResult(
        name=emp.name, region=emp.region, basic_salary=emp.basic_salary,
        actual_days=emp.actual_days, standard_days=emp.standard_days,
        allowance=emp.allowance, gross_income=round(gross_income),
        bhxh_nld=bhxh_nld, bhyt_nld=bhyt_nld, bhtn_nld=bhtn_nld, total_nld=total_nld,
        dependents=emp.dependents, family_deduction=family_deduction,
        taxable_income=taxable_income, pit=pit, net_income=net_income,
        bhxh_dn=bhxh_dn, bhyt_dn=bhyt_dn, bhtn_dn=bhtn_dn, kpcd_dn=kpcd_dn,
        total_dn_insurance=total_dn_insurance, total_dn_cost=total_dn_cost,
    )


# ── Write the result file ────────────────────────────────────────────────────

_MONEY_FMT = "#,##0"
_HEADER_FONT = Font(bold=True)
_TOTAL_FONT = Font(bold=True)

_PAYROLL_HEADERS = [
    ("STT", 5), ("Họ tên", 22), ("Vùng", 6), ("Lương cơ bản", 14),
    ("Ngày công TT/Chuẩn", 10), ("Phụ cấp", 12), ("Tổng thu nhập", 14),
    ("BHXH (NLĐ 8%)", 13), ("BHYT (NLĐ 1.5%)", 13), ("BHTN (NLĐ 1%)", 13),
    ("Tổng trích NLĐ", 13), ("Người phụ thuộc", 10), ("Giảm trừ gia cảnh", 14),
    ("Thu nhập tính thuế", 14), ("Thuế TNCN", 13), ("Lương thực nhận (Net)", 16),
]
_COST_HEADERS = [
    ("STT", 5), ("Họ tên", 22), ("Lương gộp", 14), ("BHXH DN (17.5%)", 13),
    ("BHYT DN (3%)", 12), ("BHTN DN (1%)", 12), ("KPCĐ (2%)", 12),
    ("Tổng bảo hiểm DN đóng", 15), ("Tổng chi phí DN / nhân viên", 16),
]


@log_call
def _write_header(ws, headers):
    for col, (label, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = width
    ws.freeze_panes = "A2"


@log_call
def build_payroll_workbook(results: list[EmployeeResult], out_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bảng lương"
    _write_header(ws, _PAYROLL_HEADERS)

    money_cols_payroll = (4, 6, 7, 8, 9, 10, 11, 13, 14, 15, 16)
    for i, r in enumerate(results, start=1):
        row = i + 1
        ws.append([
            i, r.name, r.region, r.basic_salary, f"{r.actual_days:g}/{r.standard_days:g}",
            r.allowance, r.gross_income, r.bhxh_nld, r.bhyt_nld, r.bhtn_nld, r.total_nld,
            r.dependents, r.family_deduction, r.taxable_income, r.pit, r.net_income,
        ])
        for col in money_cols_payroll:
            ws.cell(row=row, column=col).number_format = _MONEY_FMT

    total_row = len(results) + 2
    ws.cell(row=total_row, column=2, value="Tổng cộng").font = _TOTAL_FONT
    for col in (7, 8, 9, 10, 11, 14, 15, 16):
        letter = ws.cell(row=1, column=col).column_letter
        cell = ws.cell(row=total_row, column=col,
                        value=f"=SUM({letter}2:{letter}{total_row - 1})")
        cell.number_format = _MONEY_FMT
        cell.font = _TOTAL_FONT

    ws2 = wb.create_sheet("Chi phí DN")
    _write_header(ws2, _COST_HEADERS)
    for i, r in enumerate(results, start=1):
        row = i + 1
        ws2.append([
            i, r.name, r.gross_income, r.bhxh_dn, r.bhyt_dn, r.bhtn_dn, r.kpcd_dn,
            r.total_dn_insurance, r.total_dn_cost,
        ])
        for col in (3, 4, 5, 6, 7, 8, 9):
            ws2.cell(row=row, column=col).number_format = _MONEY_FMT

    total_row2 = len(results) + 2
    ws2.cell(row=total_row2, column=2, value="Tổng cộng").font = _TOTAL_FONT
    for col in (3, 4, 5, 6, 7, 8, 9):
        letter = ws2.cell(row=1, column=col).column_letter
        cell = ws2.cell(row=total_row2, column=col,
                        value=f"=SUM({letter}2:{letter}{total_row2 - 1})")
        cell.number_format = _MONEY_FMT
        cell.font = _TOTAL_FONT

    try:
        wb.save(out_path)
    except PermissionError:
        raise PayrollError(S.OUTPUT_FILE_LOCKED.format(name=Path(out_path).name))


# ── Parameters from the server: fetch when due, cache locally, fall back to
#    defaults ────────────────────────────────────────────────────────────
#
# Why the network call is NOT mandatory on every payroll run: this feature
# is deliberately designed to work fully offline (no AI quota spent). The
# server (`GET /rates`, see `docai-server/functions/src/index.ts`) is only
# an OPTIONAL update source — editing the rates in the Firestore Console
# takes effect on every machine within `_RATES_TTL_SECONDS`, no app
# rebuild/update needed. If the server isn't deployed yet or the machine is
# offline, it automatically falls back to the most recent fetch's cache,
# then to the hardcoded defaults in `payroll_rates.py`.

_RATES_CACHE_KEY = "payroll_rates_cache"
_RATES_TTL_SECONDS = 24 * 3600  # fast enough that a Firestore edit reaches every machine within a day


@log_call
def _parse_server_rates(raw: dict) -> PayrollRates:
    ins_raw = raw.get("insurance") or {}
    default_ins = DEFAULT_PAYROLL_RATES.insurance
    insurance = InsuranceRates(
        bhxh_nld=float(ins_raw.get("bhxh_nld", default_ins.bhxh_nld)),
        bhyt_nld=float(ins_raw.get("bhyt_nld", default_ins.bhyt_nld)),
        bhtn_nld=float(ins_raw.get("bhtn_nld", default_ins.bhtn_nld)),
        bhxh_dn=float(ins_raw.get("bhxh_dn", default_ins.bhxh_dn)),
        bhyt_dn=float(ins_raw.get("bhyt_dn", default_ins.bhyt_dn)),
        bhtn_dn=float(ins_raw.get("bhtn_dn", default_ins.bhtn_dn)),
        kpcd_dn=float(ins_raw.get("kpcd_dn", default_ins.kpcd_dn)),
    )

    vung_raw = raw.get("luong_toi_thieu_vung") or {}
    luong_toi_thieu_vung = ({str(k).upper(): int(v) for k, v in vung_raw.items()}
                            if vung_raw else dict(DEFAULT_PAYROLL_RATES.luong_toi_thieu_vung))

    brackets_raw = raw.get("tax_brackets") or []
    tax_brackets = [
        TaxBracket(float(b[0]) if b[0] is not None else float("inf"), float(b[1]))
        for b in brackets_raw
    ] if brackets_raw else list(DEFAULT_PAYROLL_RATES.tax_brackets)

    return PayrollRates(
        insurance=insurance,
        luong_co_so=int(raw.get("luong_co_so", DEFAULT_PAYROLL_RATES.luong_co_so)),
        luong_toi_thieu_vung=luong_toi_thieu_vung,
        giam_tru_ban_than=int(
            raw.get("giam_tru_ban_than", DEFAULT_PAYROLL_RATES.giam_tru_ban_than)),
        giam_tru_nguoi_phu_thuoc=int(
            raw.get("giam_tru_nguoi_phu_thuoc", DEFAULT_PAYROLL_RATES.giam_tru_nguoi_phu_thuoc)),
        tax_brackets=tax_brackets,
        version=str(raw.get("version") or "server"),
    )


@log_call
def _cached_rates(cfg: dict) -> PayrollRates | None:
    cache = cfg.get(_RATES_CACHE_KEY) or {}
    if not cache.get("rates"):
        return None
    try:
        return _parse_server_rates(cache["rates"])
    except Exception:  # noqa: BLE001 — corrupted/reshaped cache → treat as no cache
        logger.warning("Cached payroll rates are malformed — treating as no cache.",
                       exc_info=True)
        return None


@log_call
def get_active_rates(force_refresh: bool = False) -> PayrollRates:
    """The parameters currently used for payroll: prefers fetching from the
    server if the cache has expired (or `force_refresh=True`); on network/
    server error → uses the most recent cache; never successfully fetched →
    uses the app's hardcoded defaults."""
    cfg = load_config()
    cache = cfg.get(_RATES_CACHE_KEY) or {}
    stale = (time.time() - cache.get("fetched_at", 0)) > _RATES_TTL_SECONDS

    if force_refresh or stale or not cache.get("rates"):
        try:
            response = api_client.get_rates()
            raw = response.get("rates") or {}
            parsed = _parse_server_rates(raw)
            cfg[_RATES_CACHE_KEY] = {"rates": raw, "fetched_at": time.time()}
            save_config(cfg)
            return parsed
        except api_client.ApiError:
            logger.warning(
                "Could not fetch payroll rates from server — falling back to cache/defaults.",
                exc_info=True)
            # offline / server hasn't deployed the /rates route yet → fall through to cache/defaults

    cached = _cached_rates(cfg)
    return cached if cached is not None else DEFAULT_PAYROLL_RATES


# ── Main entry point ──────────────────────────────────────────────────────────

@log_call
def _staging_dir() -> str:
    out_dir = os.path.join(tempfile.gettempdir(), "DocAI", "staging")
    os.makedirs(out_dir, exist_ok=True)
    return out_dir


@log_call
def _stamp() -> str:
    return time.strftime("%Y%m%d_%H%M%S")


@dataclass
class PayrollRunResult:
    employees: list[EmployeeResult]
    out_path: str
    rates_version: str = "local-default"
    total_net: int = 0
    total_nld_insurance: int = 0
    total_dn_cost: int = 0

    def __post_init__(self):
        self.total_net = sum(e.net_income for e in self.employees)
        self.total_nld_insurance = sum(e.total_nld for e in self.employees)
        self.total_dn_cost = sum(e.total_dn_cost for e in self.employees)


@log_call
def compute_payroll_file(path: str, out_path: str | None = None,
                         rates: PayrollRates | None = None) -> PayrollRunResult:
    """Read the timesheet at `path`, compute payroll + social/health/
    unemployment insurance + personal income tax for each employee
    (deterministic, no AI call for the computation — only 1 LIGHT network
    call, no quota spent, to sync parameters if the cache is stale), write
    the result to a new Excel file in the staging folder (not the final save
    location yet — per the `modules.common.doc_set` convention). Returns a
    `PayrollRunResult`."""
    if rates is None:
        rates = get_active_rates()
    employees = read_attendance_table(path)
    results = [compute_employee(emp, rates) for emp in employees]
    if out_path is None:
        out_path = os.path.join(_staging_dir(), f"BangLuong_{_stamp()}.xlsx")
    build_payroll_workbook(results, out_path)
    return PayrollRunResult(employees=results, out_path=out_path, rates_version=rates.version)
