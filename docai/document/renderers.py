import os
import re
import threading

from PyQt6.QtGui import QPixmap, QImage

try:
    import win32com.client as _w32      # noqa: F401
    import pythoncom as _pycom          # noqa: F401
    import fitz as _fitz                # noqa: F401
    from PIL import Image as _PILImg    # noqa: F401
    HAS_PAGEPREVIEW = True
except ImportError:
    HAS_PAGEPREVIEW = False


class LazyPdfSource:
    """Nguồn trang PDF render theo yêu cầu.

    Dùng cho PDF gốc, và cho Word/PPT sau khi đã quy về 1 PDF trung gian
    qua COM. Mở file 1 lần lấy ngay tổng số trang + kích thước trang (đều
    là đọc metadata, không raster) — rẻ dù tài liệu vài nghìn trang. Chỉ
    khi `png_path(i)` được gọi thì trang đó mới thực sự được raster ra PNG
    (và cache lại), để widget xem trước chỉ tải trang nào người dùng đang
    thực sự cuộn tới thay vì raster hết cả cuốn sách.
    """

    def __init__(self, pdf_path: str, tmp_dir: str, dpi: int = 150,
                 owned_files: list[str] | None = None):
        import fitz
        self._doc = fitz.open(pdf_path)
        self.page_count = self._doc.page_count
        self._dpi = dpi
        self._tmp_dir = tmp_dir
        self._lock = threading.Lock()
        self._cache: dict[int, str] = {}
        self._owned_files = owned_files or []
        self._closed = False

    def page_size(self, index: int) -> tuple[int, int]:
        zoom = self._dpi / 72.0
        with self._lock:
            rect = self._doc[index].rect
        return max(1, int(rect.width * zoom)), max(1, int(rect.height * zoom))

    def png_path(self, index: int) -> str:
        cached = self._cache.get(index)
        if cached:
            return cached
        import fitz
        zoom = self._dpi / 72.0
        mat = fitz.Matrix(zoom, zoom)
        with self._lock:
            cached = self._cache.get(index)
            if cached:
                return cached
            page = self._doc[index]
            pix = page.get_pixmap(matrix=mat, alpha=False)
            out = os.path.join(self._tmp_dir, f"page_{index:04d}.png")
            pix.save(out)
            self._cache[index] = out
        return out

    def close(self):
        import shutil
        if self._closed:
            return
        self._closed = True
        with self._lock:
            self._doc.close()
        for f in self._owned_files:
            try:
                os.unlink(f)
            except Exception:
                pass
        shutil.rmtree(self._tmp_dir, ignore_errors=True)


class EagerPageSource:
    """Nguồn trang đã raster sẵn hết — dùng cho Excel/ảnh, vốn rất ít trang
    (số sheet, hoặc 1) nên không cần tải lười."""

    def __init__(self, png_paths: list[str], tmp_dir: str):
        self._paths = png_paths
        self.page_count = len(png_paths)
        self._tmp_dir = tmp_dir
        self._closed = False

    def page_size(self, index: int) -> tuple[int, int]:
        from PIL import Image
        with Image.open(self._paths[index]) as img:
            return img.size

    def png_path(self, index: int) -> str:
        return self._paths[index]

    def close(self):
        import shutil
        if self._closed:
            return
        self._closed = True
        shutil.rmtree(self._tmp_dir, ignore_errors=True)


def pdf_page_source(path: str, dpi: int = 150) -> LazyPdfSource:
    import tempfile
    tmp_dir = tempfile.mkdtemp(prefix="docai_pages_")
    return LazyPdfSource(os.path.abspath(path), tmp_dir, dpi)


def _word_to_pdf(path: str) -> str:
    import tempfile
    import win32com.client
    import pythoncom

    abs_path = os.path.abspath(path)
    pythoncom.CoInitialize()
    try:
        word = win32com.client.Dispatch("Word.Application")
        word.Visible = False
        doc = word.Documents.Open(abs_path)
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tf:
            pdf_path = tf.name
        doc.SaveAs2(pdf_path, FileFormat=17)
        doc.Close(False)
        word.Quit()
    finally:
        pythoncom.CoUninitialize()
    return pdf_path


def word_page_source(path: str, dpi: int = 150) -> LazyPdfSource:
    import tempfile
    pdf_path = _word_to_pdf(path)
    tmp_dir = tempfile.mkdtemp(prefix="docai_pages_")
    return LazyPdfSource(pdf_path, tmp_dir, dpi, owned_files=[pdf_path])


def _ppt_to_pdf(path: str) -> str:
    import tempfile
    import win32com.client
    import pythoncom

    abs_path = os.path.abspath(path)
    pythoncom.CoInitialize()
    try:
        ppt = win32com.client.Dispatch("PowerPoint.Application")
        pres = ppt.Presentations.Open(abs_path, ReadOnly=True, WithWindow=False)
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tf:
            pdf_path = tf.name
        pres.SaveAs(pdf_path, 32)   # ppSaveAsPDF
        pres.Close()
        ppt.Quit()
    finally:
        pythoncom.CoUninitialize()
    return pdf_path


def ppt_page_source(path: str, dpi: int = 150) -> LazyPdfSource:
    import tempfile
    pdf_path = _ppt_to_pdf(path)
    tmp_dir = tempfile.mkdtemp(prefix="docai_pages_")
    return LazyPdfSource(pdf_path, tmp_dir, dpi, owned_files=[pdf_path])


def image_page_source(path: str) -> EagerPageSource:
    import tempfile
    from PIL import Image

    tmp_dir = tempfile.mkdtemp(prefix="docai_pages_")
    out = os.path.join(tmp_dir, "page_000.png")
    Image.open(path).convert("RGB").save(out, "PNG")
    return EagerPageSource([out], tmp_dir)


def excel_page_source(path: str) -> EagerPageSource:
    png_paths, tmp_dir = _render_excel_sheets(path)
    return EagerPageSource(png_paths, tmp_dir)


def _render_excel_sheets(path: str) -> tuple[list[str], str]:
    import tempfile
    import openpyxl
    from openpyxl.utils import get_column_letter
    from PIL import Image, ImageDraw, ImageFont

    S = 1.5; CHAR_PX = 7; PT_PX = 1.33
    ROW_HDR = 38; COL_HDR = 20; DEF_CW = 64; DEF_RH = 18
    GRID_CLR = "#C0C0C0"; HDR_BG = "#E8E8E8"; HDR_FG = "#666666"
    FONT_DIR = "C:/Windows/Fonts/"
    THEME = ["#FFFFFF","#000000","#E7E6E6","#44546A","#4472C4","#ED7D31",
             "#A5A5A5","#FFC000","#5B9BD5","#70AD47"]

    _fc: dict = {}

    def _fnt(size=11, bold=False):
        key = (int(size), bold)
        if key not in _fc:
            try:
                fn = "calibrib.ttf" if bold else "calibri.ttf"
                _fc[key] = ImageFont.truetype(FONT_DIR + fn, int(size * S))
            except Exception:
                _fc[key] = ImageFont.load_default()
        return _fc[key]

    def _clr(c, dflt="white"):
        try:
            if c.type == "rgb":
                rgb = c.rgb or ""
                if len(rgb) >= 6 and rgb != "00000000":
                    return "#" + rgb[-6:]
            elif c.type == "theme":
                idx = c.theme
                if 0 <= idx < len(THEME):
                    clr = THEME[idx]
                    tint = getattr(c, "tint", 0) or 0
                    if tint > 0:
                        r, g, b = (int(clr[i:i+2], 16) for i in (1, 3, 5))
                        return f"#{int(r+(255-r)*tint):02X}{int(g+(255-g)*tint):02X}{int(b+(255-b)*tint):02X}"
                    elif tint < 0:
                        r, g, b = (int(clr[i:i+2], 16) for i in (1, 3, 5))
                        return f"#{int(r*(1+tint)):02X}{int(g*(1+tint)):02X}{int(b*(1+tint)):02X}"
                    return clr
        except Exception:
            pass
        return dflt

    def _fmt_val(cell):
        v = cell.value
        if v is None:
            return ""
        fmt = (cell.number_format or "General").strip()
        if isinstance(v, (int, float)):
            if fmt in ("General", "@", ""):
                return str(int(v)) if float(v) == int(v) else str(v)
            if "%" in fmt:
                return f"{v * 100:.1f}%"
            dec = len([c for c in fmt.split(".")[-1] if c in "0#"]) if "." in fmt else 0
            use_comma = "," in fmt
            if use_comma and dec == 0:
                return f"{int(v):,}"
            elif use_comma:
                return f"{v:,.{dec}f}"
            elif dec:
                return f"{v:.{dec}f}"
            return str(int(v)) if float(v) == int(v) else str(v)
        return str(v)

    wb = openpyxl.load_workbook(path, data_only=True)
    tmp_dir = tempfile.mkdtemp(prefix="docai_pages_")
    png_paths: list[str] = []

    for ws in wb.worksheets:
        if not ws.max_row or not ws.max_column:
            continue
        if ws.max_row == 1 and ws.max_column == 1 and not ws.cell(1, 1).value:
            continue
        MR, MC = ws.max_row, ws.max_column

        cw_list = []
        for ci in range(1, MC + 1):
            d = ws.column_dimensions.get(get_column_letter(ci))
            if d and d.hidden:     cw_list.append(0)
            elif d and d.width:    cw_list.append(max(4, int(d.width * CHAR_PX)))
            else:                  cw_list.append(DEF_CW)

        rh_list = []
        for ri in range(1, MR + 1):
            d = ws.row_dimensions.get(ri)
            if d and d.hidden:     rh_list.append(0)
            elif d and d.height:   rh_list.append(max(4, int(d.height * PT_PX)))
            else:                  rh_list.append(DEF_RH)

        merged: dict = {}
        for mrng in ws.merged_cells.ranges:
            for r in range(mrng.min_row, mrng.max_row + 1):
                for c in range(mrng.min_col, mrng.max_col + 1):
                    merged[(r, c)] = mrng

        W = int((ROW_HDR + sum(cw_list)) * S)
        H = int((COL_HDR + sum(rh_list)) * S)
        img = Image.new("RGB", (W, H), "white")
        dw = ImageDraw.Draw(img)
        rh0 = int(ROW_HDR * S)
        ch0 = int(COL_HDR * S)

        dw.rectangle([0, 0, W - 1, ch0 - 1], fill=HDR_BG)
        dw.line([(0, ch0 - 1), (W - 1, ch0 - 1)], fill=GRID_CLR)
        dw.line([(rh0 - 1, 0), (rh0 - 1, ch0 - 1)], fill=GRID_CLR)
        x = rh0
        for ci, pw in enumerate(cw_list):
            if pw == 0:
                continue
            px = int(pw * S)
            ltr = get_column_letter(ci + 1)
            bb = dw.textbbox((0, 0), ltr, font=_fnt(8))
            tw, th = bb[2]-bb[0], bb[3]-bb[1]
            dw.text((x + (px-tw)//2, (ch0-th)//2), ltr, fill=HDR_FG, font=_fnt(8))
            dw.line([(x + px - 1, 0), (x + px - 1, ch0 - 1)], fill=GRID_CLR)
            x += px

        y = ch0
        for ri, ph in enumerate(rh_list):
            if ph == 0:
                continue
            row_num = ri + 1
            rh_px = int(ph * S)
            dw.rectangle([0, y, rh0 - 1, y + rh_px - 1], fill=HDR_BG)
            lbl = str(row_num)
            bb = dw.textbbox((0, 0), lbl, font=_fnt(8))
            tw2, th2 = bb[2]-bb[0], bb[3]-bb[1]
            dw.text((rh0//2 - tw2//2, y + rh_px//2 - th2//2), lbl, fill=HDR_FG, font=_fnt(8))
            dw.line([(0, y + rh_px - 1), (rh0 - 1, y + rh_px - 1)], fill=GRID_CLR)
            dw.line([(rh0 - 1, y), (rh0 - 1, y + rh_px - 1)], fill=GRID_CLR)

            x = rh0
            for ci, pw in enumerate(cw_list):
                if pw == 0:
                    continue
                col_num = ci + 1
                cw_px = int(pw * S)
                cell = ws.cell(row_num, col_num)
                mrng = merged.get((row_num, col_num))
                slave = mrng and (mrng.min_row != row_num or mrng.min_col != col_num)

                if not slave:
                    dw2 = sum(int(cw_list[c-1]*S) for c in range(mrng.min_col, mrng.max_col+1)) if mrng else cw_px
                    dh2 = sum(int(rh_list[r-1]*S) for r in range(mrng.min_row, mrng.max_row+1)) if mrng else rh_px
                    bg = "white"
                    try:
                        if cell.fill and cell.fill.patternType == "solid":
                            bg = _clr(cell.fill.fgColor, "white")
                    except Exception:
                        pass
                    dw.rectangle([x, y, x+dw2-1, y+dh2-1], fill=bg)
                    text = _fmt_val(cell)
                    if text:
                        try:
                            bold   = bool(cell.font.bold) if cell.font else False
                            fsize  = float(cell.font.size or 11) if cell.font else 11
                            fg     = _clr(cell.font.color, "#000000") if cell.font else "#000000"
                            halign = (cell.alignment.horizontal or "general") if cell.alignment else "general"
                        except Exception:
                            bold, fsize, fg, halign = False, 11, "#000000", "general"
                        fnt = _fnt(fsize, bold)
                        pad = int(3 * S)
                        try:
                            while len(text) > 1:
                                bb = dw.textbbox((0, 0), text, font=fnt)
                                if bb[2]-bb[0] <= dw2 - 2*pad:
                                    break
                                text = text[:-1]
                            bb = dw.textbbox((0, 0), text, font=fnt)
                            tw3 = bb[2]-bb[0]
                            th3 = bb[3]-bb[1]
                        except Exception:
                            tw3 = len(text)*7
                            th3 = 13
                        if halign in ("center", "centerContinuous", "distributed"):
                            tx = x + (dw2 - tw3) // 2
                        elif halign == "right":
                            tx = x + dw2 - tw3 - pad
                        else:
                            tx = x + pad
                        dw.text((tx, y + (dh2 - th3) // 2), text, fill=fg, font=fnt)

                    def _bl(side, x1, y1, x2, y2, _cell=cell):
                        try:
                            bs = getattr(_cell.border, side, None)
                            if bs and bs.border_style and bs.border_style != "none":
                                clr = _clr(bs.color, "#000000") if bs.color else "#000000"
                                width = 2 if bs.border_style in ("medium", "thick") else 1
                                dw.line([(x1, y1), (x2, y2)], fill=clr, width=int(width*S))
                        except Exception:
                            pass
                    _bl("top",    x,        y,        x+dw2-1, y)
                    _bl("bottom", x,        y+dh2-1,  x+dw2-1, y+dh2-1)
                    _bl("left",   x,        y,        x,        y+dh2-1)
                    _bl("right",  x+dw2-1,  y,        x+dw2-1, y+dh2-1)

                dw.line([(x, y+rh_px-1), (x+cw_px-1, y+rh_px-1)], fill=GRID_CLR)
                dw.line([(x+cw_px-1, y), (x+cw_px-1, y+rh_px-1)], fill=GRID_CLR)
                x += cw_px
            y += rh_px

        safe = re.sub(r'[^\w]', '_', ws.title)
        out = os.path.join(tmp_dir, f"sheet_{safe}.png")
        img.save(out, "PNG")
        png_paths.append(out)

    return png_paths, tmp_dir


def pil_to_qimage(pil_img) -> QImage:
    """An toàn gọi từ luồng nền (khác QPixmap, QImage không cần luồng GUI).
    `.copy()` để QImage sở hữu buffer riêng, tách khỏi bytes Python cục bộ
    trước khi vượt qua ranh giới luồng."""
    pil_img = pil_img.convert("RGB")
    data = pil_img.tobytes("raw", "RGB")
    qimg = QImage(data, pil_img.width, pil_img.height,
                  pil_img.width * 3, QImage.Format.Format_RGB888)
    return qimg.copy()


def pil_to_qpixmap(pil_img) -> QPixmap:
    return QPixmap.fromImage(pil_to_qimage(pil_img))
