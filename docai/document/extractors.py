import csv
import io
import os

from .models import AttachedFile

try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    import openpyxl
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False


def extract_word(path: str) -> AttachedFile:
    doc = DocxDocument(path)
    lines = []
    body = doc.element.body
    for child in body:
        tag = child.tag.split("}")[-1] if "}" in child.tag else child.tag
        if tag == "p":
            from docx.text.paragraph import Paragraph
            p = Paragraph(child, doc)
            lines.append(p.text)
        elif tag == "tbl":
            from docx.table import Table
            t = Table(child, doc)
            for row in t.rows:
                cells = [c.text.replace("\n", " ").strip() for c in row.cells]
                lines.append(" | ".join(cells))
            lines.append("")
    return AttachedFile(path=path, name=os.path.basename(path),
                        file_type="word", claude_content="\n".join(lines))


def extract_excel(path: str) -> AttachedFile:
    wb = openpyxl.load_workbook(path, data_only=True)
    parts = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = [[str(c) if c is not None else "" for c in row]
                for row in ws.iter_rows(values_only=True)]
        while rows and all(c == "" for c in rows[-1]):
            rows.pop()
        parts.append(f"--- Sheet: {sname} ---")
        buf = io.StringIO()
        csv.writer(buf).writerows(rows)
        parts.extend([buf.getvalue().rstrip(), ""])
    return AttachedFile(path=path, name=os.path.basename(path),
                        file_type="excel", claude_content="\n".join(parts))
