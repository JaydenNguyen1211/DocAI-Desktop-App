"""Áp lệnh sửa của AI lên file Word/Excel ngay tại chỗ (ghi đè file gốc).

Server (Claude) quyết định danh sách lệnh sửa; module này chỉ thực thi.
  · Word  — bộ thao tác đầy đủ trong `word_ops` (thêm/xóa đoạn, xóa trang,
            chèn vào vị trí cụ thể, định dạng, tìm & thay…).
  · Excel — bản này mới có thêm dòng mới ở cuối sheet.
"""
from pathlib import Path

try:
    from docx import Document as DocxDocument   # noqa: F401 — chỉ để kiểm tra có lib
except ImportError:
    DocxDocument = None  # type: ignore

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore

from . import word_ops
from .word_ops import WordOpError

# Chỉ định dạng mở (OOXML) mới sửa được — .doc/.xls cũ thì không.
EDITABLE_EXT = {".docx": "word", ".xlsx": "excel"}

EXCEL_OPS = ("append_row",)


class EditError(Exception):
    """Không sửa được file — thông báo tiếng Việt cho người dùng."""


def check_editable(path: str, file_type: str | None) -> None:
    """Ném EditError nếu file không nằm trong phạm vi sửa được."""
    if file_type not in ("word", "excel"):
        raise EditError("Bản này chỉ sửa được file Word và Excel.")

    suffix = Path(path).suffix.lower()
    if suffix not in EDITABLE_EXT:
        raise EditError(
            f"Chưa sửa được định dạng {suffix} — hãy lưu lại thành "
            f"{'.docx' if file_type == 'word' else '.xlsx'} rồi mở lại."
        )
    if file_type == "word" and DocxDocument is None:
        raise EditError("Thiếu thư viện python-docx để sửa file Word.")
    if file_type == "excel" and openpyxl is None:
        raise EditError("Thiếu thư viện openpyxl để sửa file Excel.")


def document_outline(path: str, file_type: str | None) -> str:
    """Mô tả cấu trúc file để gửi lên server làm ngữ cảnh cho Claude."""
    try:
        if file_type == "word":
            return word_ops.outline_text(path)
        if file_type == "excel":
            return _excel_outline(path)
    except EditError:
        raise
    except Exception as exc:  # noqa: BLE001 — file hỏng/không đọc được
        raise EditError(f"Không đọc được cấu trúc file: {exc}")
    return ""


def apply_edits(path: str, file_type: str | None, edits: list[dict]) -> list[str]:
    """Thực thi danh sách lệnh sửa. Trả về mô tả từng lệnh (để báo trong chat)."""
    check_editable(path, file_type)
    if not edits:
        return []

    try:
        if file_type == "word":
            return word_ops.apply_edits(path, edits)
        return _apply_excel(path, edits)
    except WordOpError as exc:
        raise EditError(str(exc))


# ── Excel ──────────────────────────────────────────────────────────────────

def _excel_outline(path: str) -> str:
    wb = openpyxl.load_workbook(path, data_only=True)
    lines = []
    for ws in wb.worksheets:
        lines.append(f"Sheet «{ws.title}»: {ws.max_row} dòng × {ws.max_column} cột.")
    return "\n".join(lines)


def _apply_excel(path: str, edits: list[dict]) -> list[str]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active if wb.active is not None else wb.worksheets[0]
    notes = []

    for edit in edits:
        op = (edit.get("op") or "").strip()
        if op not in EXCEL_OPS:
            raise EditError(
                f"Với file Excel, bản này mới hỗ trợ thêm dòng mới "
                f"(nhận được «{op}»).")
        text = (edit.get("text") or "").strip()
        if not text:
            raise EditError("Không rõ cần thêm nội dung gì vào dòng mới.")
        ws.append([text])
        notes.append(f"thêm «{text}» vào dòng mới")

    try:
        wb.save(path)
    except PermissionError:
        raise EditError(
            f"Không ghi được «{Path(path).name}» — file đang mở trong Excel, "
            "hãy đóng lại rồi thử lại.")
    return notes
